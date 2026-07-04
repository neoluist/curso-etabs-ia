# -*- coding: utf-8 -*-
"""Genera una hoja Excel PRESENTABLE del espectro de respuesta E.030-2026 con
FORMULAS VIVAS (no valores hardcodeados) + 2 graficos (X-X / Y-Y), modelada en la
hoja del usuario (ESPECTRO DE RESPUESTA SISMICA - 2026.xlsx).

Diseno: parametros de entrada en celdas (Z,U,S,Tp,Tl,g, R0/Ia/Ip por direccion);
el factor C (con rama corta E.030) y Sa = Z*U*C*S*g/R se calculan por FORMULA en cada
fila de la tabla de periodos, referenciando esas celdas. Cambiar un parametro recalcula
todo el espectro y los graficos. Hoja extra "DATOS TXT" para el From File de ETABS.

Lo usa el servidor (POST /espectro/excel) y tambien corre solo para pruebas:
    python espectro_excel.py            -> escribe espectro_demo.xlsx
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import get_column_letter

# Periodos T (s) estandar de la hoja E.030 (48 puntos). Es el eje de periodos.
PERIODOS_DEF = [0, 0.02, 0.04, 0.06, 0.08, 0.1, 0.12, 0.14, 0.16, 0.18, 0.2,
                0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75,
                0.8, 0.85, 0.9, 0.95, 1, 1.1, 1.2, 1.3, 1.4, 1.5, 1.6, 1.7,
                1.8, 1.9, 2, 2.25, 2.5, 2.75, 3, 4, 5, 6, 7, 8, 9, 10]

# Paleta (estilo paper/ingenieria)
AZUL = "1F4E79"        # encabezados / titulo
AZUL_CLARO = "D6E4F0"  # fondo de seccion
AMBAR = "C55A11"       # acentos
GRIS = "808080"
BORDE = "BFBFBF"
INPUT_FONT = "0000FF"  # convencion: entradas en AZUL
FUENTE = "Calibri"

_thin = Side(style="thin", color=BORDE)
_box = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _f(**kw):
    kw.setdefault("name", FUENTE)
    return Font(**kw)


def build_espectro_xlsx(params: dict) -> bytes:
    """params: {z,u,s,tp,tl,g, r0x,iax,ipx, r0y,iay,ipy, periodos?,
                zona?,suelo?,uso?,sistemaX?,sistemaY?,proyecto?}"""
    p = params or {}
    periodos = [float(t) for t in (p.get("periodos") or PERIODOS_DEF)]
    g = float(p.get("g", 9.81))

    wb = Workbook()
    ws = wb.active
    ws.title = "ESPECTRO E030-2026"
    ws.sheet_view.showGridLines = False

    # ---- Anchos de columna ----
    for col, w in {"A": 16, "B": 11, "C": 11, "D": 3,
                   "E": 9, "F": 9, "G": 12, "H": 12}.items():
        ws.column_dimensions[col].width = w

    def put(coord, value, *, font=None, fill=None, align=None, fmt=None, border=False):
        c = ws[coord]
        c.value = value
        if font:
            c.font = font
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if align:
            c.alignment = align
        if fmt:
            c.number_format = fmt
        if border:
            c.border = _box
        return c

    cen = Alignment(horizontal="center", vertical="center")
    izq = Alignment(horizontal="left", vertical="center")
    der = Alignment(horizontal="right", vertical="center")

    # ===== TITULO =====
    ws.merge_cells("A1:H1")
    put("A1", "ESPECTRO DE RESPUESTA SISMICA - E.030 (2026)",
        font=_f(bold=True, size=15, color="FFFFFF"), fill=AZUL, align=cen)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:H2")
    sub = p.get("proyecto") or "Diseno sismorresistente E.030 - Sa = Z*U*C*S*g / R"
    put("A2", sub, font=_f(italic=True, size=9, color=GRIS), align=cen)

    # ===== PARAMETROS (entradas en AZUL) =====
    put("A4", "PARAMETROS", font=_f(bold=True, size=10, color=AZUL))
    filas_simple = [
        ("A5", "Z (zona)", "B5", p.get("z"), p.get("zona")),
        ("A6", "U (uso)", "B6", p.get("u"), p.get("uso")),
        ("A7", "S (suelo)", "B7", p.get("s"), p.get("suelo")),
        ("A8", "Tp (s)", "B8", p.get("tp"), None),
        ("A9", "Tl (s)", "B9", p.get("tl"), None),
        ("A10", "g (m/s2)", "B10", g, None),
    ]
    for lab_c, lab, val_c, val, nota in filas_simple:
        put(lab_c, lab, font=_f(size=9), align=izq, border=True)
        put(val_c, val, font=_f(size=9, color=INPUT_FONT, bold=True), align=cen,
            fmt="0.###", border=True)
        if nota:
            put("C" + lab_c[1:], nota, font=_f(size=8, italic=True, color=GRIS), align=izq)

    # Encabezado X-X / Y-Y
    put("A12", "Por direccion", font=_f(bold=True, size=9, color=AZUL))
    put("B12", "X-X", font=_f(bold=True, size=9, color="FFFFFF"), fill=AZUL, align=cen, border=True)
    put("C12", "Y-Y", font=_f(bold=True, size=9, color=AMBAR), fill=AZUL_CLARO, align=cen, border=True)
    if p.get("sistemaX"):
        put("D12", p.get("sistemaX"), font=_f(size=8, italic=True, color=GRIS))
    par_dir = [
        ("A13", "R0 (sistema)", "B13", p.get("r0x"), "C13", p.get("r0y")),
        ("A14", "Ia (altura)", "B14", p.get("iax"), "C14", p.get("iay")),
        ("A15", "Ip (planta)", "B15", p.get("ipx"), "C15", p.get("ipy")),
    ]
    for lab_c, lab, bx, vx, cy, vy in par_dir:
        put(lab_c, lab, font=_f(size=9), align=izq, border=True)
        put(bx, vx, font=_f(size=9, color=INPUT_FONT, bold=True), align=cen, fmt="0.###", border=True)
        put(cy, vy, font=_f(size=9, color=INPUT_FONT, bold=True), align=cen, fmt="0.###", border=True)
    # R = R0*Ia*Ip por FORMULA (negro = calculo)
    put("A16", "R = R0*Ia*Ip", font=_f(size=9, bold=True), align=izq, border=True)
    put("B16", "=B13*B14*B15", font=_f(size=9, bold=True), align=cen, fmt="0.###", border=True)
    put("C16", "=C13*C14*C15", font=_f(size=9, bold=True), align=cen, fmt="0.###", border=True)

    # ===== TABLA DEL ESPECTRO (E:H) =====
    r0 = 4  # fila del encabezado de la tabla
    heads = [("E", "T (s)"), ("F", "C"), ("G", "Sa X (m/s2)"), ("H", "Sa Y (m/s2)")]
    for col, txt in heads:
        put(f"{col}{r0}", txt, font=_f(bold=True, size=9, color="FFFFFF"), fill=AZUL,
            align=cen, border=True)
    n = len(periodos)
    first = r0 + 1
    last = r0 + n
    for i, t in enumerate(periodos):
        r = first + i
        put(f"E{r}", t, font=_f(size=9, color=INPUT_FONT), align=cen, fmt="0.00", border=True)
        # Factor C E.030 (con rama corta 1+7.5*T/Tp); Tp=$B$8, Tl=$B$9
        put(f"F{r}",
            f"=IF(E{r}<0.2*$B$8,1+7.5*E{r}/$B$8,IF(E{r}<=$B$8,2.5,"
            f"IF(E{r}<$B$9,2.5*$B$8/E{r},2.5*$B$8*$B$9/E{r}^2)))",
            font=_f(size=9), align=cen, fmt="0.000", border=True)
        # Sa = Z*U*C*S*g / R   (Z=$B$5,U=$B$6,S=$B$7,g=$B$10, Rx=$B$16, Ry=$C$16)
        put(f"G{r}", f"=$B$5*$B$6*F{r}*$B$7*$B$10/$B$16",
            font=_f(size=9), align=cen, fmt="0.0000", border=True)
        put(f"H{r}", f"=$B$5*$B$6*F{r}*$B$7*$B$10/$C$16",
            font=_f(size=9), align=cen, fmt="0.0000", border=True)

    # ===== GRAFICO Sa - T (X-X y Y-Y juntos) =====
    chart = ScatterChart()
    chart.title = "Espectro de pseudo-aceleraciones  Sa - T"
    chart.style = 2
    chart.height = 9.5
    chart.width = 18
    chart.x_axis.title = "Periodo  T (s)"
    chart.y_axis.title = "Sa (m/s2)"
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    xref = Reference(ws, min_col=5, min_row=first, max_row=last)  # T
    for col, nombre, color in ((7, "Sa X-X", "1F4E79"), (8, "Sa Y-Y", "C55A11")):
        yref = Reference(ws, min_col=col, min_row=r0, max_row=last)  # incluye encabezado (titulo de serie)
        s = Series(yref, xref, title_from_data=True)
        s.marker.symbol = "none"
        s.smooth = False
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = 22000  # EMU ~ 1.7pt
        chart.series.append(s)
    ws.add_chart(chart, "E20" if n < 16 else f"E{last + 3}")

    # ===== HOJA "DATOS TXT" (From File de ETABS: T  Sa, separado X/Y) =====
    ws2 = wb.create_sheet("DATOS TXT")
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 12
    put2 = lambda coord, val, **k: _put_plain(ws2, coord, val, **k)
    put2("A1", "T (X-X)", bold=True, fill=AZUL, color="FFFFFF")
    put2("B1", "Sa (X-X)", bold=True, fill=AZUL, color="FFFFFF")
    put2("D1", "T (Y-Y)", bold=True, fill=AMBAR, color="FFFFFF")
    put2("E1", "Sa (Y-Y)", bold=True, fill=AMBAR, color="FFFFFF")
    for i in range(n):
        r = 2 + i
        src = first + i
        ws2[f"A{r}"] = f"=+'ESPECTRO E030-2026'!E{src}"
        ws2[f"B{r}"] = f"=+'ESPECTRO E030-2026'!G{src}"
        ws2[f"D{r}"] = f"=+'ESPECTRO E030-2026'!E{src}"
        ws2[f"E{r}"] = f"=+'ESPECTRO E030-2026'!H{src}"
        for cc in ("A", "B", "D", "E"):
            ws2[f"{cc}{r}"].number_format = "0.0000"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _put_plain(ws, coord, val, *, bold=False, fill=None, color=None):
    c = ws[coord]
    c.value = val
    c.font = Font(name=FUENTE, bold=bold, color=color or "000000", size=9)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c


if __name__ == "__main__":
    demo = {
        "z": 0.45, "u": 1.0, "s": 1.05, "tp": 0.6, "tl": 2.0, "g": 9.81,
        "r0x": 8, "iax": 1.0, "ipx": 1.0, "r0y": 6, "iay": 0.75, "ipy": 1.0,
        "zona": "Z4", "suelo": "S1", "uso": "C (comun)",
        "sistemaX": "Portico de concreto", "sistemaY": "Muros de concreto",
        "proyecto": "Espectro E.030-2026 (demo)",
    }
    data = build_espectro_xlsx(demo)
    with open("espectro_demo.xlsx", "wb") as fh:
        fh.write(data)
    print("escrito espectro_demo.xlsx", len(data), "bytes")
